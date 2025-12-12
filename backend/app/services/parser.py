import os
import re
import json
from groq import Groq
from pathlib import Path

class DocumentParser:
    def __init__(self):
        self.client = Groq(api_key=os.getenv("GROQ_API_KEY"))
    
    async def parse_document_smart(self, file_path: str, file_type: str) -> dict:
        """Парсит документ с Groq (с fallback на regex)"""
        try:
            # Извлеки текст
            text = self._extract_text(file_path, file_type)
            if not text:
                return {"positions": [], "metadata": {"confidence": 0, "method": "error"}}
            
            # Парсь через Llama
            return await self.parse_with_groq(text)
        except Exception as e:
            print(f"Groq error: {e}, falling back to regex")
            text = self._extract_text(file_path, file_type)
            return self._parse_with_regex(text)
    
    async def parse_with_groq(self, text: str) -> dict:
        """Парсинг через Groq Llama 3.1"""
        prompt = f"""Анализируй этот текст из закупочной заявки.
Извлеки позиции (товары/услуги) в формате JSON.

ОБЯЗАТЕЛЬНО вернуть JSON:
{{
  "positions": [
    {{"pos": 1, "name": "Наименование", "unit": "м", "qty": 140}}
  ]
}}

ТЕКСТ:
{text[:3000]}

ТОЛЬКО JSON, БЕЗ КОММЕНТАРИЕВ!"""
        
        try:
            response = self.client.messages.create(
                model="llama-3.1-70b-versatile",
                messages=[{"role": "user", "content": prompt}],
                temperature=0.1,
                max_tokens=2048,
            )
            
            result = self._extract_json(response.content)
            if result.get("positions"):
                result["metadata"] = {"confidence": 0.92, "method": "groq"}
                return result
        except Exception as e:
            print(f"Groq parse error: {e}")
        
        # Fallback
        return self._parse_with_regex(text)
    
    def _extract_text(self, file_path: str, file_type: str) -> str:
        """Извлекает текст из документа"""
        try:
            if file_type == "docx":
                from docx import Document
                doc = Document(file_path)
                return "\n".join([p.text for p in doc.paragraphs])
            
            elif file_type == "pdf":
                import PyPDF2
                with open(file_path, "rb") as f:
                    reader = PyPDF2.PdfReader(f)
                    return "\n".join([page.extract_text() or "" for page in reader.pages])
            
            elif file_type == "xlsx":
                import openpyxl
                wb = openpyxl.load_workbook(file_path)
                text = ""
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    for row in ws.iter_rows(values_only=True):
                        text += " | ".join([str(c) if c else "" for c in row]) + "\n"
                return text
        except Exception as e:
            print(f"Extract error: {e}")
        
        return ""
    
    def _parse_with_regex(self, text: str) -> dict:
        """Fallback на regex"""
        positions = []
        lines = text.split("\n")
        
        for line in lines:
            # С трубой: "1 | название | ед | кол"
            match = re.match(r'^(\d+)\s*\|\s*(.*?)\s*\|\s*([а-яА-ЯмШТ]+)\s*\|\s*(\d+)', line)
            
            # Без трубы: "1 название м 140"
            if not match:
                match = re.match(r'^(\d+)\s+(.*?)\s+([а-яА-ЯмШТ]+)\s+(\d+)$', line)
            
            if match:
                positions.append({
                    "pos": int(match.group(1)),
                    "name": match.group(2).strip(),
                    "unit": match.group(3).strip(),
                    "qty": int(match.group(4)),
                })
        
        return {
            "positions": positions,
            "metadata": {"confidence": 0.6, "method": "regex"}
        }
    
    def _extract_json(self, text: str) -> dict:
        """Безопасный парс JSON из ответа"""
        try:
            match = re.search(r'\{.*\}', text, re.DOTALL)
            if match:
                return json.loads(match.group())
        except:
            pass
        return {"positions": []}
